from openpyxl import load_workbook
wb = load_workbook('Ekta Indoor Events.xlsx', data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print('===', sheet, '===')
    print('rows', len(rows)-1, 'cols', len(header))
    print('header:', header)
    if sheet == 'Fun':
        games = {}
        for row in rows[1:]:
            if not row or all(c is None for c in row): continue
            game = row[7]
            if game is None: continue
            games[game] = games.get(game,0)+1
        print('games counts:', games)
    elif sheet == 'Singles':
        fields = ['Carrom Singles','Chess','Pool Table','Squash Singles','Table Tennis']
        counts={f:0 for f in fields}
        entries=[]
        for row in rows[1:]:
            if not row or all(c is None for c in row): continue
            for i,f in enumerate(fields, start=8):
                if i < len(row) and row[i] is not None:
                    counts[f]+=1
                    entries.append((row[1],row[2],f,row[i]))
        print('event counts:', counts)
        dupnames={}
        for r in rows[1:]:
            if not r or all(c is None for c in r): continue
            name=r[1]
            if name:
                dupnames[name]=dupnames.get(name,0)+1
        dups=[(n,c) for n,c in dupnames.items() if c>1]
        print('duplicate names count', len(dups), 'sample', dups[:10])
    elif sheet == 'Doubles':
        fields = ['Competitive Doubles TT','Name and Phone number of your Partner for Carrom','Competitive Games Doubles','Name and Phone number of your Table Tennis Partner']
        counts={f:0 for f in fields}
        for row in rows[1:]:
            if not row or all(c is None for c in row): continue
            for i,f in enumerate(fields, start=8):
                if i < len(row) and row[i] is not None:
                    counts[f]+=1
        print('doubles counts:', counts)
        names=[row[1] for row in rows[1:] if row and row[1]]
        print('unique names', len(set(names)), 'total entries', len(names))
    print()
