from openpyxl import load_workbook
wb = load_workbook('Ekta_Schedule_With_Matchups.xlsx', data_only=True)
ws = wb['Squash']
rows = list(ws.iter_rows(values_only=True))
print('rows', len(rows), 'cols', ws.max_column)
for i, row in enumerate(rows[:80]):
    print(i, row)
