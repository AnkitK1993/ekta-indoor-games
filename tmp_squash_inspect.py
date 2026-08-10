from openpyxl import load_workbook
wb = load_workbook('Ekta Indoor Events.xlsx', data_only=True)
ws = wb['Singles']
rows = list(ws.iter_rows(values_only=True))
# Find Squash section boundaries
start = None
end = None
for i,row in enumerate(rows):
    if row and row[0] == 'SQUASH TOURNAMENT - PLAYER GROUPS':
        start = i
    if start is not None and row and row[0] == 'Time' and row[1] == 'Match' and row[2] == 'Category':
        start = i
        break
print('start index', start)
for idx in range(max(0,start-5), min(len(rows), start+40)):
    print(idx, rows[idx])
