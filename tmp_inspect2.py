from openpyxl import load_workbook
wb = load_workbook('Ekta_Schedule_With_Matchups.xlsx', data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # count non-empty rows and first row
    rows = list(ws.iter_rows(values_only=True))
    nonempty = [r for r in rows if any(cell is not None for cell in r)]
    print('===', sheet, '===')
    print('rows', ws.max_row, 'cols', ws.max_column, 'nonempty', len(nonempty))
    print('header', rows[0])
    # find match rows based on presence of labels and vs
    match_rows = [r for r in rows if r and any(str(c).strip().startswith('M') for c in r if c)]
    print('first 5 match rows sample:')
    for r in match_rows[:5]:
        print(' ', r)
    # if sheet has structured pairings, count A/B rows
    if ws.max_column >= 7:
        player_cells = [r[4] for r in match_rows if len(r) > 4 and r[4] is not None]
        print('sample player A values:', player_cells[:3])
    print()
