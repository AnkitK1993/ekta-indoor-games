from openpyxl import load_workbook
import re
wb = load_workbook('Ekta_Schedule_With_Matchups.xlsx', data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    match_rows=[]
    byes=[]
    placeholders=[]
    for row in rows:
        if not row or all(cell is None for cell in row):
            continue
        if len(row) >= 2 and isinstance(row[1], str) and re.match(r'^(M|GA|GB|GA-|GB-|GA|GB|G|Quarterfinal|Semifinal|Final|Round|—|Time)', row[1], re.I):
            if row[1].strip().lower() in ('match','round','time','s.no','player name','game / category','tournament date:'):
                continue
        if len(row) >= 7 and row[4] is not None and row[5] == 'vs':
            match_rows.append(row)
            a=str(row[4])
            b=str(row[6]) if len(row) > 6 and row[6] is not None else ''
            if 'sits out' in str(row[3]).lower() or 'advances' in a.lower() or 'advances' in b.lower() or 'rested' in a.lower() or 'rested' in b.lower():
                byes.append(row)
            if 'Winner' in a or 'Winner' in b:
                placeholders.append(row)
    print('===', sheet, '===')
    print('total rows', len(rows), 'nonempty', sum(1 for r in rows if r and any(c is not None for c in r)))
    print('matches found', len(match_rows), 'byes/odd rows', len(byes), 'placeholders', len(placeholders))
    print('sample match rows:')
    for r in match_rows[:5]:
        print(' ', r)
    print()
