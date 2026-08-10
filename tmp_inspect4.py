from openpyxl import load_workbook
wb = load_workbook('Ekta Indoor Events.xlsx', data_only=True)
print('sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print('\n===', sheet, '=== rows', ws.max_row, 'cols', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
        print(row)
